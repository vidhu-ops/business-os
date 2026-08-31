"""Import legacy IIDATECH users from an Excel export into the CRM user store.

Always restores password_hash from the spreadsheet so original passwords keep working.

Usage:
  python tools/import_legacy_users_xlsx.py "c:\\Users\\weewee\\Downloads\\user (1).xlsx"
"""
from __future__ import annotations

import json
import sys
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))


def _parse_created(raw: object) -> str:
    if raw is None:
        return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
    s = str(raw).strip().strip('"')
    if not s:
        return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
    try:
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        return dt.astimezone(timezone.utc).isoformat().replace("+00:00", "Z")
    except Exception:
        return s


def _truthy(raw: object) -> bool:
    return str(raw or "").strip().lower() in {"1", "true", "yes", "y"}


def rows_from_xlsx(xlsx: Path) -> list[dict]:
    import openpyxl

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [str(h or "").strip() for h in rows[0]]
    out = []
    for raw in rows[1:]:
        out.append({headers[i]: raw[i] if i < len(raw) else None for i in range(len(headers))})
    return out


def record_from_row(row: dict, existing: dict | None = None) -> dict:
    email = str(row.get("email") or "").strip().lower()
    username = str(row.get("username") or "").strip() or email.split("@")[0]
    credits = 0
    try:
        credits = max(0, int(row.get("credits") or 0))
    except (TypeError, ValueError):
        credits = 0
    pw = str(row.get("password_hash") or "").strip()
    record = dict(existing or {})
    record["email"] = email
    if not record.get("name"):
        record["name"] = username.replace(".", " ").replace("_", " ").strip() or email.split("@")[0]
    record["username"] = username
    record.setdefault("created_at", _parse_created(row.get("created_at")))
    # Always restore original spreadsheet password so legacy logins keep working.
    if pw:
        record["password_hash"] = pw
    try:
        current_rem = int(record.get("credits_remaining") or 0)
    except (TypeError, ValueError):
        current_rem = 0
    rem = max(current_rem, credits)
    record["credits_remaining"] = rem
    try:
        current_total = int(record.get("credits_total") or 0)
    except (TypeError, ValueError):
        current_total = 0
    record["credits_total"] = max(current_total, rem, credits)
    record["source"] = "legacy_iidatech_users_xlsx"
    record["legacy_flags"] = {
        "ai_create_access_paid": _truthy(row.get("ai_create_access_paid")),
        "is_subscriber": _truthy(row.get("is_subscriber")),
        "financial_tools_access_paid": _truthy(row.get("financial_tools_access_paid")),
        "event_management_access_paid": _truthy(row.get("event_management_access_paid")),
        "ai_create_expiry": str(row.get("ai_create_expiry") or "").strip().strip('"') or None,
        "financial_tools_expiry": str(row.get("financial_tools_expiry") or "").strip().strip('"') or None,
        "event_management_expiry": str(row.get("event_management_expiry") or "").strip().strip('"') or None,
        "subscription_expiry": str(row.get("subscription_expiry") or "").strip().strip('"') or None,
        "pm_access_expiry": str(row.get("pm_access_expiry") or "").strip().strip('"') or None,
        "legacy_id": row.get("id"),
    }
    if not record.get("plan"):
        record["plan"] = "free"
    record["imported_at"] = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
    return record


def main() -> int:
    xlsx = Path(sys.argv[1] if len(sys.argv) > 1 else r"c:\Users\weewee\Downloads\user (1).xlsx")
    if not xlsx.is_file():
        print(f"File not found: {xlsx}")
        return 1

    from backend.services.user_store import load_users, save_users

    rows = rows_from_xlsx(xlsx)
    users = load_users()
    created = updated = skipped = 0
    seed: dict = {}

    for row in rows:
        email = str(row.get("email") or "").strip().lower()
        if not email or "@" not in email:
            skipped += 1
            continue
        existing = users.get(email) if isinstance(users.get(email), dict) else None
        record = record_from_row(row, existing)
        users[email] = record
        seed[email] = {
            "email": email,
            "name": record.get("name"),
            "username": record.get("username"),
            "password_hash": record.get("password_hash"),
            "credits_remaining": record.get("credits_remaining"),
            "credits_total": record.get("credits_total"),
            "created_at": record.get("created_at"),
            "plan": record.get("plan", "free"),
            "source": "legacy_iidatech_users_xlsx",
            "legacy_flags": record.get("legacy_flags"),
            "imported_at": record.get("imported_at"),
        }
        if existing:
            updated += 1
        else:
            created += 1

    save_users(users)
    seed_blob = json.dumps(seed, indent=2, ensure_ascii=False, default=str)
    seed_paths = [
        ROOT / "backend" / "data" / "legacy_iidatech_users_seed.json",
        ROOT / "business_build_outputs" / "legacy_iidatech_users_seed.json",
    ]
    written: list[str] = []
    for seed_path in seed_paths:
        seed_path.parent.mkdir(parents=True, exist_ok=True)
        seed_path.write_text(seed_blob, encoding="utf-8")
        written.append(str(seed_path))
    print(
        json.dumps(
            {
                "created": created,
                "updated": updated,
                "skipped": skipped,
                "total_users": len(users),
                "seed_users": len(seed),
                "seeds": written,
            },
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
