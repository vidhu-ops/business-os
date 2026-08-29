"""Parse CSV / XLSX lead sheets into normalized row dicts."""
from __future__ import annotations

import csv
import io
from typing import Any

HEADER_MAP = {
    "email": "email",
    "e-mail": "email",
    "e_mail": "email",
    "mail": "email",
    "name": "name",
    "full name": "name",
    "fullname": "name",
    "contact": "name",
    "contact name": "name",
    "phone": "phone",
    "mobile": "phone",
    "whatsapp": "phone",
    "tel": "phone",
    "company": "company",
    "organization": "company",
    "organisation": "company",
    "org": "company",
    "source": "source",
    "channel": "source",
    "campaign": "source",
    "city": "city",
    "location": "place",
    "place": "place",
    "country": "country",
    "notes": "notes",
    "note": "notes",
    "comment": "notes",
    "website": "website",
    "url": "website",
}


def _norm_header(value: str) -> str:
    return " ".join(str(value or "").strip().lower().replace("_", " ").split())


def _map_row(headers: list[str], values: list[Any]) -> dict[str, str]:
    out: dict[str, str] = {}
    for idx, header in enumerate(headers):
        key = HEADER_MAP.get(_norm_header(header))
        if not key:
            continue
        raw = values[idx] if idx < len(values) else ""
        text = str(raw or "").strip()
        if text:
            out[key] = text[:240]
    if out.get("city") and not out.get("place"):
        out["place"] = out["city"]
    return out


def parse_lead_sheet(content: bytes, filename: str = "") -> list[dict[str, str]]:
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xls"):
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise ValueError("Excel import needs openpyxl on the server") from exc
        book = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = book.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(cell or "") for cell in rows[0]]
        out = []
        for row in rows[1:]:
            mapped = _map_row(headers, list(row or []))
            if mapped:
                out.append(mapped)
        return out

    text = content.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
    except Exception:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    rows = list(reader)
    if not rows:
        return []
    headers = [str(cell or "") for cell in rows[0]]
    out = []
    for row in rows[1:]:
        mapped = _map_row(headers, list(row or []))
        if mapped:
            out.append(mapped)
    return out
