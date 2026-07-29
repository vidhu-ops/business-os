"""Load CRM Automation market verification workbook (IIDATECH proprietary reference)."""
from __future__ import annotations

import re
from functools import lru_cache
from pathlib import Path
from typing import Any

_DEFAULT_PATH = (
    Path(__file__).resolve().parents[1] / "proprietary_data" / "crm_automation_market_verification.xlsx"
)

_CRM_TOPIC_RE = re.compile(
    r"\b(crm|customer relationship|sales automation|lead management|pipeline|hubspot|pipedrive|freshsales|zoho crm|revops)\b",
    re.I,
)

_DATA_SHEETS = (
    "Tier1_Government_Data",
    "Tier2_Company_Disclosures",
    "Tier3_Syndicated_Reports",
    "Tier4_Trade_Associations_India",
    "Tier5_Pricing_Benchmarks",
    "Tier6_Industry_Vertical_Data",
)


def workbook_path(custom: str | Path | None = None) -> Path:
    if custom:
        return Path(custom)
    env = __import__("os").getenv("CRM_VERIFICATION_WORKBOOK_PATH", "").strip()
    if env:
        return Path(env)
    return _DEFAULT_PATH


def is_crm_workbook_topic(topic: str, industry: str = "") -> bool:
    blob = f"{topic} {industry}".strip()
    if not blob:
        return False
    if _CRM_TOPIC_RE.search(blob):
        return True
    try:
        try:
            from streamlit_app import classify_topic_domain
        except ImportError:
            from app import classify_topic_domain

        return classify_topic_domain(topic, industry) == "crm_automation"
    except Exception:
        return False


def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _is_verified(status: str) -> bool:
    s = str(status or "").upper()
    return s.startswith("VERIFIED")


def _header_map(headers: list[str]) -> dict[str, int]:
    out: dict[str, int] = {}
    for i, h in enumerate(headers):
        key = _cell_str(h).lower()
        if key:
            out[key] = i
    return out


def _pick(row: list[Any], hdr: dict[str, int], *names: str) -> str:
    for name in names:
        idx = hdr.get(name.lower())
        if idx is not None and idx < len(row):
            v = _cell_str(row[idx])
            if v:
                return v
    return ""


@lru_cache(maxsize=4)
def _load_workbook_records_cached(path_str: str, mtime_ns: int) -> tuple[dict[str, Any], ...]:
    p = Path(path_str)
    if not p.is_file():
        return ()
    try:
        import openpyxl
    except ImportError:
        return ()

    records: list[dict[str, Any]] = []
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    try:
        for sheet_name in _DATA_SHEETS:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            rows = ws.iter_rows(values_only=True)
            header_row = next(rows, None)
            if not header_row:
                continue
            headers = [_cell_str(h) for h in header_row]
            hdr = _header_map(headers)
            if not hdr:
                continue
            for row in rows:
                if not row or not any(_cell_str(c) for c in row):
                    continue
                cells = list(row)
                first = _cell_str(cells[0]) if cells else ""
                if first.lower().startswith("cross-check"):
                    continue
                data_point = _pick(
                    cells,
                    hdr,
                    "data point",
                    "vertical / segment",
                    "report title",
                    "vendor",
                    "company",
                    "publisher",
                )
                value = _pick(
                    cells,
                    hdr,
                    "value",
                    "reported 2026 market size",
                    "reported share or size",
                    "value / headline",
                    "entry/starter tier",
                    "total revenue",
                )
                if not data_point and not value:
                    if first and not first.lower().startswith("cross-check"):
                        data_point = first
                if not data_point:
                    continue
                geo = _pick(cells, hdr, "geography", "geography ")
                entity = _pick(cells, hdr, "company", "vendor", "publisher")
                verified_raw = _pick(cells, hdr, "verified?", "verified")
                rec = {
                    "sheet": sheet_name,
                    "category": sheet_name.replace("_", " "),
                    "geography": geo or "Global",
                    "entity": entity,
                    "data_point": data_point,
                    "value": value,
                    "publisher": _pick(cells, hdr, "publisher", "dataset / report name", "filing type", "report title"),
                    "url": _pick(cells, hdr, "url", "url (sec edgar / ir)"),
                    "verified": verified_raw,
                    "verified_bool": _is_verified(verified_raw),
                    "notes": _pick(cells, hdr, "notes / caveats", "notes"),
                    "date": _pick(cells, hdr, "date published", "date", "fiscal year", "date/age"),
                    "vertical": _pick(cells, hdr, "vertical / segment"),
                    "growth_note": _pick(cells, hdr, "fastest-growing vertical (per this publisher)"),
                }
                records.append(rec)
    finally:
        wb.close()
    return tuple(records)


def load_workbook_records(path: str | Path | None = None) -> tuple[dict[str, Any], ...]:
    p = workbook_path(path)
    mtime_ns = int(p.stat().st_mtime_ns) if p.is_file() else 0
    return _load_workbook_records_cached(str(p.resolve()), mtime_ns)


def records_for_geography(geography: str, *, path: str | None = None) -> list[dict[str, Any]]:
    geo = str(geography or "Global").strip().lower()
    out: list[dict[str, Any]] = []
    for rec in load_workbook_records(path):
        rg = str(rec.get("geography") or "").lower()
        if not rg or rg == "global" or geo in rg or rg in geo:
            out.append(dict(rec))
        elif geo in {"global", "worldwide"}:
            out.append(dict(rec))
        elif "india" in geo and "india" in rg:
            out.append(dict(rec))
        elif geo in {"us", "usa", "united states"} and rg in {"us", "usa", "united states"}:
            out.append(dict(rec))
    return out


def verified_records(geography: str = "Global", *, path: str | None = None) -> list[dict[str, Any]]:
    return [r for r in records_for_geography(geography, path=path) if r.get("verified_bool")]


def search_records(query: str, geography: str = "Global", *, limit: int = 20) -> list[dict[str, Any]]:
    q = str(query or "").lower()
    if not q:
        return []
    scored: list[tuple[int, dict[str, Any]]] = []
    for rec in records_for_geography(geography):
        blob = " ".join(
            str(rec.get(k) or "")
            for k in ("data_point", "entity", "value", "publisher", "notes", "vertical", "growth_note")
        ).lower()
        hits = sum(1 for token in q.split() if len(token) > 2 and token in blob)
        if hits:
            scored.append((hits, rec))
    scored.sort(key=lambda x: (-x[0], not x[1].get("verified_bool")))
    return [dict(r) for _, r in scored[:limit]]
